from __future__ import annotations

from typing import Optional

from ..vision_tool import GeminiVision, load_prompt

PROMPT_PATH = "packages/prompts/vision_extract_rich.md"


def extract_xlsx(content: bytes, *, vision: Optional[GeminiVision], limits: dict) -> str:
    try:
        import io
        from openpyxl import load_workbook
    except Exception:
        return ""

    prompt = load_prompt(PROMPT_PATH)
    max_imgs = int(limits.get("XLSX_VISION_MAX_IMAGES", 10))
    max_lines = int(limits.get("XLSX_MAX_CELL_LINES", 5000))

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    parts = []

    for ws in wb.worksheets:
        # Cell text (bounded)
        lines = []
        try:
            for row in ws.iter_rows(values_only=True):
                vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if vals:
                    lines.append(" | ".join(vals))
                if len(lines) >= max_lines:
                    break
        except Exception:
            pass

        if lines:
            suffix = ""
            if len(lines) >= max_lines:
                suffix = f"\n[TRUNCATED] sheet_lines>={max_lines}"
            parts.append(f"\n\n--- SHEET {ws.title} CELLS ---\n" + "\n".join(lines) + suffix)

        # Embedded images (bounded)
        if vision and vision.enabled() and max_imgs > 0:
            imgs = getattr(ws, "_images", None) or []
            for idx, img in enumerate(imgs[:max_imgs]):
                try:
                    img_bytes = img._data()
                    vis = vision.analyze_image(prompt=prompt, image_bytes=img_bytes, mime="image/png")
                    if vis:
                        parts.append(f"\n--- SHEET {ws.title} IMAGE {idx+1} VISION ---\n{vis}")
                except Exception:
                    continue

    return "\n".join(parts).strip()